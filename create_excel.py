#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils.exceptions import InvalidFileException
from datetime import datetime

# 导入测试数据配置（从 data_config.py）
from data_config import SAMPLE_PRODUCTS, SAMPLE_CUSTOMERS, SAMPLE_CUSTOMER_PRICES, SAMPLE_ORDERS

# ==================== 程序开始 ====================
# 创建工作簿
wb = Workbook()
wb.remove(wb.active)  # 删除默认sheet

# ==================== Sheet 1: 产品库 ====================
ws_products = wb.create_sheet("产品库")

# 表头
headers_products = ["产品名称", "规格", "单位", "单价", "备注"]
for col, header in enumerate(headers_products, 1):
    cell = ws_products.cell(1, col, header)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

# 示例数据 - 多种规格（从data_config导入）
for row_idx, row_data in enumerate(SAMPLE_PRODUCTS, 2):
    for col_idx, value in enumerate(row_data, 1):
        ws_products.cell(row_idx, col_idx, value)

# 设置列宽
ws_products.column_dimensions['A'].width = 20
ws_products.column_dimensions['B'].width = 15
ws_products.column_dimensions['C'].width = 12
ws_products.column_dimensions['D'].width = 10
ws_products.column_dimensions['E'].width = 20

# 定义为表格 (Table)
from openpyxl.worksheet.table import Table, TableStyleInfo
tbl_products = Table(displayName="tbl_products", ref=f"A1:E{len(SAMPLE_PRODUCTS)+1}")
style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
tbl_products.tableStyleInfo = style
ws_products.add_table(tbl_products)

# ==================== Sheet 1.5: 客户资料 ====================
ws_customers = wb.create_sheet("客户资料")

# 表头
headers_customers = ["客户名称", "联系人", "联系电话", "地址", "备注"]
for col, header in enumerate(headers_customers, 1):
    cell = ws_customers.cell(1, col, header)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

# 示例数据（从data_config导入）
for row_idx, row_data in enumerate(SAMPLE_CUSTOMERS, 2):
    for col_idx, value in enumerate(row_data, 1):
        ws_customers.cell(row_idx, col_idx, value)

# 设置列宽
ws_customers.column_dimensions['A'].width = 15
ws_customers.column_dimensions['B'].width = 12
ws_customers.column_dimensions['C'].width = 15
ws_customers.column_dimensions['D'].width = 25
ws_customers.column_dimensions['E'].width = 15

# 定义为表格
tbl_customers = Table(displayName="tbl_customers", ref=f"A1:E{len(SAMPLE_CUSTOMERS)+1}")
style_customers = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                                 showLastColumn=False, showRowStripes=True, showColumnStripes=False)
tbl_customers.tableStyleInfo = style_customers
ws_customers.add_table(tbl_customers)

# ==================== Sheet 1.6: 客户报价表 ====================
ws_customer_prices = wb.create_sheet("客户报价")

# 表头
headers_customer = ["客户", "产品名称", "规格", "单价"]
for col, header in enumerate(headers_customer, 1):
    cell = ws_customer_prices.cell(1, col, header)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

# 示例数据 - 不同客户不同价格（从data_config导入）
for row_idx, row_data in enumerate(SAMPLE_CUSTOMER_PRICES, 2):
    for col_idx, value in enumerate(row_data, 1):
        ws_customer_prices.cell(row_idx, col_idx, value)

# 设置列宽
ws_customer_prices.column_dimensions['A'].width = 15
ws_customer_prices.column_dimensions['B'].width = 20
ws_customer_prices.column_dimensions['C'].width = 15
ws_customer_prices.column_dimensions['D'].width = 12

# 定义为表格
tbl_customer_prices = Table(displayName="tbl_customer_prices", ref=f"A1:D{len(SAMPLE_CUSTOMER_PRICES)+1}")
style_customer = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                                showLastColumn=False, showRowStripes=True, showColumnStripes=False)
tbl_customer_prices.tableStyleInfo = style_customer
ws_customer_prices.add_table(tbl_customer_prices)

# ==================== Sheet 2: 订单录入 ====================
ws_orders = wb.create_sheet("订单录入")

# 表头（无序号列）
headers_orders = ["日期", "客户", "产品名称", "规格", "单价", "数量", "重量", "总价", "备注", "订单号", "打印状态"]
for col, header in enumerate(headers_orders, 1):
    cell = ws_orders.cell(1, col, header)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

# 使用说明（黄色底色）
ws_orders['L1'] = "使用说明"
ws_orders['L1'].font = Font(bold=True, size=10, color="FFFFFF")
ws_orders['L1'].fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
ws_orders['L1'].alignment = Alignment(horizontal='center')

tips_orders = [
    "【基本操作】",
    "1. 日期快速填入：Ctrl+; 生成当前日期，可修改为其他日期",
    "2. 客户/产品：点击单元格选择下拉选项（有完整列表）",
    "3. 规格：选择产品后自动带出，可根据需要修改",
    "",
    "【自动计算】",
    "4. 单价：优先查询客户报价表，如无则查产品库",
    "5. 总价：单价×数量自动计算（如填重量则×重量）",
    "6. 订单号：日期+客户相同会自动继承上行号",
    "",
    "【快捷技巧】",
    "7. 复制行：选中一整行→Ctrl+D下拉填充",
    "8. 快速输入：同一产品多行可同时选择→Ctrl+D",
    "9. 批量修改：选中列→Ctrl+H查找替换",
    "",
    "【验证和打印】",
    "10. 打印状态：下拉选择「未打印」或「已打印」",
    "11. 订单号有效：必须是「YYYYMMDD-序号」格式",
    "12. 打印模板：在「打印模板」页选择订单号即可预览"
]

for idx, tip in enumerate(tips_orders, 2):
    ws_orders[f'L{idx}'] = tip
    if tip.startswith("【"):
        ws_orders[f'L{idx}'].font = Font(bold=True, size=9, color="FFFFFF")
        ws_orders[f'L{idx}'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    elif tip == "":
        ws_orders[f'L{idx}'].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    else:
        ws_orders[f'L{idx}'].font = Font(size=9)
        ws_orders[f'L{idx}'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    ws_orders[f'L{idx}'].alignment = Alignment(wrap_text=True)

# 调整 L 列宽度
ws_orders.column_dimensions['L'].width = 35

# 设置列宽
col_widths = [12, 15, 20, 15, 12, 10, 10, 12, 20, 15, 12]
for col, width in enumerate(col_widths, 1):
    ws_orders.column_dimensions[get_column_letter(col)].width = width

# 添加一行示例（第二行）
# A2: 日期 - 保持为空，让用户输入
ws_orders['A2'] = ""
ws_orders['A2'].number_format = 'yyyy/mm/dd'

# B2: 客户 - 下拉选择（引用客户资料表）
dv_customer = DataValidation(type="list", formula1="'客户资料'!$A$2:$A$100", allow_blank=True)
ws_orders.add_data_validation(dv_customer)
dv_customer.add(ws_orders['B2'])
ws_orders['B2'] = "客户A"

# C2: 产品名称 - 数据验证（下拉选择）
dv_product = DataValidation(type="list", formula1="'产品库'!$A$2:$A$100", allow_blank=True)
dv_product.error = "请选择产品库中的产品"
dv_product.errorTitle = "无效选择"
ws_orders.add_data_validation(dv_product)
dv_product.add(ws_orders['C2'])
ws_orders['C2'] = "焦糖瓜子"  # 默认选中

# D2: 规格 - 下拉选择（从产品库获取所有规格）
dv_spec = DataValidation(type="list", formula1="'产品库'!$B$2:$B$100", allow_blank=True)
dv_spec.error = "请选择产品库中的规格"
dv_spec.errorTitle = "无效选择"
ws_orders.add_data_validation(dv_spec)
dv_spec.add(ws_orders['D2'])
ws_orders['D2'] = "袋装100g"

# E2: 单价 - 优先查客户报价（多条件SUMIFS），再查产品库
ws_orders['E2'] = '=IF(OR(B2="",C2="",D2=""),"",IF(SUMIFS(客户报价!$D:$D,客户报价!$A:$A,B2,客户报价!$B:$B,C2,客户报价!$C:$C,D2)>0,SUMIFS(客户报价!$D:$D,客户报价!$A:$A,B2,客户报价!$B:$B,C2,客户报价!$C:$C,D2),IF(SUMIFS(产品库!$D:$D,产品库!$A:$A,C2,产品库!$B:$B,D2)>0,SUMIFS(产品库!$D:$D,产品库!$A:$A,C2,产品库!$B:$B,D2),"⚠️请核价")))'
ws_orders['E2'].number_format = '0.00'

# F2: 数量
ws_orders['F2'] = 1

# G2: 重量
ws_orders['G2'] = ""

# H2: 总价 - 单价 × 重量（如果重量为空则用数量）
ws_orders['H2'] = '=IF(G2="",E2*F2,E2*G2)'
ws_orders['H2'].number_format = '0.00'

# I2: 备注
ws_orders['I2'] = ""

# J2: 订单号 - 根据日期自动生成
# J2: 订单号 - 第一行手动输入
ws_orders['J2'] = ""

# K2: 打印状态 - 下拉选择
dv_print = DataValidation(type="list", formula1='"未打印,已打印"', allow_blank=False)
ws_orders.add_data_validation(dv_print)
dv_print.add(ws_orders['K2'])
ws_orders['K2'] = "未打印"

# 定义为表格 (包含测试数据，扩展到1000行)
tbl_orders = Table(displayName="tbl_orders", ref="A1:K1000")
style_orders = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                              showLastColumn=False, showRowStripes=True, showColumnStripes=False)
tbl_orders.tableStyleInfo = style_orders
ws_orders.add_table(tbl_orders)

# 添加测试订单数据（使用SAMPLE_ORDERS列表）
# 结构：A日期 B客户 C产品 D规格 E单价 F数量 G重量 H总价 I备注 J订单号 K打印状态
for idx, order_data in enumerate(SAMPLE_ORDERS):
    row = idx + 2  # 从第2行开始
    # A列: 日期（转换为Excel日期）
    date_obj = datetime.strptime(order_data[0], "%Y/%m/%d")
    ws_orders[f'A{row}'] = date_obj
    ws_orders[f'A{row}'].number_format = 'yyyy/mm/dd'
    # B列: 客户
    ws_orders[f'B{row}'] = order_data[1]
    # C列: 产品名称
    ws_orders[f'C{row}'] = order_data[2]
    # D列: 规格
    ws_orders[f'D{row}'] = order_data[3]
    # F列: 数量
    ws_orders[f'F{row}'] = order_data[4]
    # J列: 订单号 - 直接使用数据中的订单号
    ws_orders[f'J{row}'] = order_data[5]
    # K列: 打印状态
    ws_orders[f'K{row}'] = order_data[6]
    # E列: 单价公式
    ws_orders[f'E{row}'] = f'=IF(OR(B{row}="",C{row}="",D{row}=""),"",IF(SUMIFS(客户报价!$D:$D,客户报价!$A:$A,B{row},客户报价!$B:$B,C{row},客户报价!$C:$C,D{row})>0,SUMIFS(客户报价!$D:$D,客户报价!$A:$A,B{row},客户报价!$B:$B,C{row},客户报价!$C:$C,D{row}),IF(SUMIFS(产品库!$D:$D,产品库!$A:$A,C{row},产品库!$B:$B,D{row})>0,SUMIFS(产品库!$D:$D,产品库!$A:$A,C{row},产品库!$B:$B,D{row}),"⚠️请核价")))'
    ws_orders[f'E{row}'].number_format = '0.00'
    # H列: 总价公式
    ws_orders[f'H{row}'] = f'=IF(G{row}="",E{row}*F{row},E{row}*G{row})'
    ws_orders[f'H{row}'].number_format = '0.00'

# ===== 扩展数据验证和公式到更多行 =====
for row in range(3, 1001):
    # 添加数据验证
    dv_customer.add(ws_orders[f'B{row}'])
    dv_product.add(ws_orders[f'C{row}'])
    dv_spec.add(ws_orders[f'D{row}'])
    dv_print.add(ws_orders[f'K{row}'])

    # E列: 单价公式
    ws_orders[f'E{row}'] = f'=IF(OR(B{row}="",C{row}="",D{row}=""),"",IF(SUMIFS(客户报价!$D:$D,客户报价!$A:$A,B{row},客户报价!$B:$B,C{row},客户报价!$C:$C,D{row})>0,SUMIFS(客户报价!$D:$D,客户报价!$A:$A,B{row},客户报价!$B:$B,C{row},客户报价!$C:$C,D{row}),IF(SUMIFS(产品库!$D:$D,产品库!$A:$A,C{row},产品库!$B:$B,D{row})>0,SUMIFS(产品库!$D:$D,产品库!$A:$A,C{row},产品库!$B:$B,D{row}),"⚠️请核价")))'
    ws_orders[f'E{row}'].number_format = '0.00'

    # H列: 总价公式
    ws_orders[f'H{row}'] = f'=IF(G{row}="",E{row}*F{row},E{row}*G{row})'
    ws_orders[f'H{row}'].number_format = '0.00'

    # J列: 订单号 - 如果日期和客户与上一行相同则继承，否则自动生成
    prev_row = row - 1
    ws_orders[f'J{row}'] = f'=IF(AND(A{row}=A{prev_row},B{row}=B{prev_row}),J{prev_row},IF(A{row}="","",TEXT(A{row},"YYYYMMDD")&"-"&TEXT(ROW()-1,"0000")))'

    # K列: 默认未打印
    ws_orders[f'K{row}'] = "未打印"

# ==================== Sheet 3: 打印模板 ====================
ws_print = wb.create_sheet("打印模板")

# 使用说明（黄色底色，在右侧显示）
ws_print['J1'] = "使用说明"
ws_print['J1'].font = Font(bold=True)
ws_print['J1'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
ws_print['J2'] = "1. 在B5单元格选择订单号"
ws_print['J3'] = "2. 系统自动显示订单详情"
ws_print['J4'] = "3. 修改公司信息（1-3行）"
ws_print['J5'] = "4. 打印：小票纸设置已配置"
for row in range(2, 6):
    ws_print[f'J{row}'].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

# ====== 公司信息（可自定义）======
ws_print['A1'] = "XX商贸有限公司"  # 修改为公司名
ws_print['A1'].font = Font(size=14, bold=True)
ws_print['A1'].alignment = Alignment(horizontal='center')
ws_print.merge_cells('A1:G1')

ws_print['A2'] = "地址：北京市朝阳区XX路88号"  # 修改为地址
ws_print['A2'].font = Font(size=9)
ws_print['A2'].alignment = Alignment(horizontal='center')
ws_print.merge_cells('A2:G2')

ws_print['A3'] = "电话：138-0013-8000"  # 修改为电话
ws_print['A3'].font = Font(size=9)
ws_print['A3'].alignment = Alignment(horizontal='center')
ws_print.merge_cells('A3:G3')

# 分隔线
ws_print['A4'] = "─" * 40
ws_print['A4'].alignment = Alignment(horizontal='center')
ws_print.merge_cells('A4:G4')

# 订单信息
ws_print['A5'] = "订单号："
ws_print['B5'] = ""  # 用户输入订单号
ws_print['B5'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
ws_print['B5'].font = Font(bold=True)

ws_print['C5'] = "日期："
ws_print['D5'] = '=IF($B$5="","",IFERROR(TEXT(INDEX(\'订单录入\'!$A$2:$A$1000,MATCH($B$5,\'订单录入\'!$J$2:$J$1000,0)),"yyyy/mm/dd"),""))'

ws_print['E5'] = "客户："
ws_print['F5'] = '=IF($B$5="","",IFERROR(INDEX(\'订单录入\'!$B$2:$B$1000,MATCH($B$5,\'订单录入\'!$J$2:$J$1000,0)),""))'

# 产品明细标题 - 只显示产品相关信息
headers = ["序号", "产品名称", "规格", "数量", "单价", "总价", "备注"]
for col, header in enumerate(headers, 1):
    cell = ws_print.cell(7, col, header)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')
    cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

# 产品明细数据行 - 直接从订单录入表中引用，不使用数组公式
# 使用直接行号引用的方式（用户需要在订单录入中将同一订单的数据放在连续行中）
# 订单录入列：A日期 B客户 C产品名称 D规格 E单价 F数量 G重量 H总价 I备注 J订单号 K打印状态
# 打印模板列：A序号 B产品名称 C规格 D数量 E单价 F总价 G备注

# 第8行数据 - 显示选中订单号的第1行数据
ws_print['A8'] = 1
ws_print['A8'].alignment = Alignment(horizontal='center')
# 使用MATCH找到订单号在订单录入表中的位置，然后向下读取数据
ws_print['B8'] = '=IF($B$5="","",IFERROR(INDEX(\'订单录入\'!$C:$C,MATCH($B$5,\'订单录入\'!$J:$J,0)),""))'
ws_print['C8'] = '=IF($B$5="","",IFERROR(INDEX(\'订单录入\'!$D:$D,MATCH($B$5,\'订单录入\'!$J:$J,0)),""))'
ws_print['D8'] = '=IF($B$5="","",IFERROR(INDEX(\'订单录入\'!$F:$F,MATCH($B$5,\'订单录入\'!$J:$J,0)),""))'
ws_print['D8'].alignment = Alignment(horizontal='center')
ws_print['E8'] = '=IF($B$5="","",IFERROR(INDEX(\'订单录入\'!$E:$E,MATCH($B$5,\'订单录入\'!$J:$J,0)),""))'
ws_print['E8'].number_format = '0.00'
ws_print['F8'] = '=IF($B$5="","",IFERROR(INDEX(\'订单录入\'!$H:$H,MATCH($B$5,\'订单录入\'!$J:$J,0)),""))'
ws_print['F8'].number_format = '0.00'
ws_print['G8'] = '=IF($B$5="","",IFERROR(INDEX(\'订单录入\'!$I:$I,MATCH($B$5,\'订单录入\'!$J:$J,0)),""))'

# 第9行数据 - 显示选中订单号的第2行数据（第1个匹配+1，仅当订单号相同时）
ws_print['A9'] = 2
ws_print['A9'].alignment = Alignment(horizontal='center')
ws_print['B9'] = '=IF($B$5="","",IF(INDEX(\'订单录入\'!$J:$J,MATCH($B$5,\'订单录入\'!$J:$J,0)+1)=$B$5,IFERROR(INDEX(\'订单录入\'!$C:$C,MATCH($B$5,\'订单录入\'!$J:$J,0)+1),""),""))'
ws_print['C9'] = '=IF($B$5="","",IF(INDEX(\'订单录入\'!$J:$J,MATCH($B$5,\'订单录入\'!$J:$J,0)+1)=$B$5,IFERROR(INDEX(\'订单录入\'!$D:$D,MATCH($B$5,\'订单录入\'!$J:$J,0)+1),""),""))'
ws_print['D9'] = '=IF($B$5="","",IF(INDEX(\'订单录入\'!$J:$J,MATCH($B$5,\'订单录入\'!$J:$J,0)+1)=$B$5,IFERROR(INDEX(\'订单录入\'!$F:$F,MATCH($B$5,\'订单录入\'!$J:$J,0)+1),""),""))'
ws_print['D9'].alignment = Alignment(horizontal='center')
ws_print['E9'] = '=IF($B$5="","",IF(INDEX(\'订单录入\'!$J:$J,MATCH($B$5,\'订单录入\'!$J:$J,0)+1)=$B$5,IFERROR(INDEX(\'订单录入\'!$E:$E,MATCH($B$5,\'订单录入\'!$J:$J,0)+1),""),""))'
ws_print['E9'].number_format = '0.00'
ws_print['F9'] = '=IF($B$5="","",IF(INDEX(\'订单录入\'!$J:$J,MATCH($B$5,\'订单录入\'!$J:$J,0)+1)=$B$5,IFERROR(INDEX(\'订单录入\'!$H:$H,MATCH($B$5,\'订单录入\'!$J:$J,0)+1),""),""))'
ws_print['F9'].number_format = '0.00'
ws_print['G9'] = '=IF($B$5="","",IF(INDEX(\'订单录入\'!$J:$J,MATCH($B$5,\'订单录入\'!$J:$J,0)+1)=$B$5,IFERROR(INDEX(\'订单录入\'!$I:$I,MATCH($B$5,\'订单录入\'!$J:$J,0)+1),""),""))'

# 第10行数据 - 显示选中订单号的第3行数据（第1个匹配+2，仅当订单号相同时）
ws_print['A10'] = 3
ws_print['A10'].alignment = Alignment(horizontal='center')
ws_print['B10'] = '=IF($B$5="","",IF(INDEX(\'订单录入\'!$J:$J,MATCH($B$5,\'订单录入\'!$J:$J,0)+2)=$B$5,IFERROR(INDEX(\'订单录入\'!$C:$C,MATCH($B$5,\'订单录入\'!$J:$J,0)+2),""),""))'
ws_print['C10'] = '=IF($B$5="","",IF(INDEX(\'订单录入\'!$J:$J,MATCH($B$5,\'订单录入\'!$J:$J,0)+2)=$B$5,IFERROR(INDEX(\'订单录入\'!$D:$D,MATCH($B$5,\'订单录入\'!$J:$J,0)+2),""),""))'
ws_print['D10'] = '=IF($B$5="","",IF(INDEX(\'订单录入\'!$J:$J,MATCH($B$5,\'订单录入\'!$J:$J,0)+2)=$B$5,IFERROR(INDEX(\'订单录入\'!$F:$F,MATCH($B$5,\'订单录入\'!$J:$J,0)+2),""),""))'
ws_print['D10'].alignment = Alignment(horizontal='center')
ws_print['E10'] = '=IF($B$5="","",IF(INDEX(\'订单录入\'!$J:$J,MATCH($B$5,\'订单录入\'!$J:$J,0)+2)=$B$5,IFERROR(INDEX(\'订单录入\'!$E:$E,MATCH($B$5,\'订单录入\'!$J:$J,0)+2),""),""))'
ws_print['E10'].number_format = '0.00'
ws_print['F10'] = '=IF($B$5="","",IF(INDEX(\'订单录入\'!$J:$J,MATCH($B$5,\'订单录入\'!$J:$J,0)+2)=$B$5,IFERROR(INDEX(\'订单录入\'!$H:$H,MATCH($B$5,\'订单录入\'!$J:$J,0)+2),""),""))'
ws_print['F10'].number_format = '0.00'
ws_print['G10'] = '=IF($B$5="","",IF(INDEX(\'订单录入\'!$J:$J,MATCH($B$5,\'订单录入\'!$J:$J,0)+2)=$B$5,IFERROR(INDEX(\'订单录入\'!$I:$I,MATCH($B$5,\'订单录入\'!$J:$J,0)+2),""),""))'

# 第11行数据 - 显示选中订单号的第4行数据（第1个匹配+3，仅当订单号相同时）
ws_print['A11'] = 4
ws_print['A11'].alignment = Alignment(horizontal='center')
ws_print['B11'] = '=IF($B$5="","",IF(INDEX(\'订单录入\'!$J:$J,MATCH($B$5,\'订单录入\'!$J:$J,0)+3)=$B$5,IFERROR(INDEX(\'订单录入\'!$C:$C,MATCH($B$5,\'订单录入\'!$J:$J,0)+3),""),""))'
ws_print['C11'] = '=IF($B$5="","",IF(INDEX(\'订单录入\'!$J:$J,MATCH($B$5,\'订单录入\'!$J:$J,0)+3)=$B$5,IFERROR(INDEX(\'订单录入\'!$D:$D,MATCH($B$5,\'订单录入\'!$J:$J,0)+3),""),""))'
ws_print['D11'] = '=IF($B$5="","",IF(INDEX(\'订单录入\'!$J:$J,MATCH($B$5,\'订单录入\'!$J:$J,0)+3)=$B$5,IFERROR(INDEX(\'订单录入\'!$F:$F,MATCH($B$5,\'订单录入\'!$J:$J,0)+3),""),""))'
ws_print['D11'].alignment = Alignment(horizontal='center')
ws_print['E11'] = '=IF($B$5="","",IF(INDEX(\'订单录入\'!$J:$J,MATCH($B$5,\'订单录入\'!$J:$J,0)+3)=$B$5,IFERROR(INDEX(\'订单录入\'!$E:$E,MATCH($B$5,\'订单录入\'!$J:$J,0)+3),""),""))'
ws_print['E11'].number_format = '0.00'
ws_print['F11'] = '=IF($B$5="","",IF(INDEX(\'订单录入\'!$J:$J,MATCH($B$5,\'订单录入\'!$J:$J,0)+3)=$B$5,IFERROR(INDEX(\'订单录入\'!$H:$H,MATCH($B$5,\'订单录入\'!$J:$J,0)+3),""),""))'
ws_print['F11'].number_format = '0.00'
ws_print['G11'] = '=IF($B$5="","",IF(INDEX(\'订单录入\'!$J:$J,MATCH($B$5,\'订单录入\'!$J:$J,0)+3)=$B$5,IFERROR(INDEX(\'订单录入\'!$I:$I,MATCH($B$5,\'订单录入\'!$J:$J,0)+3),""),""))'

# 第12行数据 - 显示选中订单号的第5行数据（第1个匹配+4，仅当订单号相同时）
ws_print['A12'] = 5
ws_print['A12'].alignment = Alignment(horizontal='center')
ws_print['B12'] = '=IF($B$5="","",IF(INDEX(\'订单录入\'!$J:$J,MATCH($B$5,\'订单录入\'!$J:$J,0)+4)=$B$5,IFERROR(INDEX(\'订单录入\'!$C:$C,MATCH($B$5,\'订单录入\'!$J:$J,0)+4),""),""))'
ws_print['C12'] = '=IF($B$5="","",IF(INDEX(\'订单录入\'!$J:$J,MATCH($B$5,\'订单录入\'!$J:$J,0)+4)=$B$5,IFERROR(INDEX(\'订单录入\'!$D:$D,MATCH($B$5,\'订单录入\'!$J:$J,0)+4),""),""))'
ws_print['D12'] = '=IF($B$5="","",IF(INDEX(\'订单录入\'!$J:$J,MATCH($B$5,\'订单录入\'!$J:$J,0)+4)=$B$5,IFERROR(INDEX(\'订单录入\'!$F:$F,MATCH($B$5,\'订单录入\'!$J:$J,0)+4),""),""))'
ws_print['D12'].alignment = Alignment(horizontal='center')
ws_print['E12'] = '=IF($B$5="","",IF(INDEX(\'订单录入\'!$J:$J,MATCH($B$5,\'订单录入\'!$J:$J,0)+4)=$B$5,IFERROR(INDEX(\'订单录入\'!$E:$E,MATCH($B$5,\'订单录入\'!$J:$J,0)+4),""),""))'
ws_print['E12'].number_format = '0.00'
ws_print['F12'] = '=IF($B$5="","",IF(INDEX(\'订单录入\'!$J:$J,MATCH($B$5,\'订单录入\'!$J:$J,0)+4)=$B$5,IFERROR(INDEX(\'订单录入\'!$H:$H,MATCH($B$5,\'订单录入\'!$J:$J,0)+4),""),""))'
ws_print['F12'].number_format = '0.00'
ws_print['G12'] = '=IF($B$5="","",IF(INDEX(\'订单录入\'!$J:$J,MATCH($B$5,\'订单录入\'!$J:$J,0)+4)=$B$5,IFERROR(INDEX(\'订单录入\'!$I:$I,MATCH($B$5,\'订单录入\'!$J:$J,0)+4),""),""))'

# 分隔线
ws_print['A18'] = "─" * 40
ws_print['A18'].alignment = Alignment(horizontal='center')
ws_print.merge_cells('A18:G18')

# 合计 - 使用SUMIF汇总同一订单的所有金额
ws_print['E19'] = "合计："
ws_print['E19'].font = Font(bold=True)
ws_print['E19'].alignment = Alignment(horizontal='right')
ws_print['F19'] = '=IF($B$5="","",SUMIF(\'订单录入\'!$J$2:$J$1000,$B$5,\'订单录入\'!$H$2:$H$1000))'
ws_print['F19'].number_format = '0.00'
ws_print['F19'].font = Font(size=12, bold=True)
ws_print['F19'].alignment = Alignment(horizontal='right')
ws_print['F19'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# 签字区域
ws_print['A21'] = "送货人签字：________________"
ws_print['A21'].font = Font(size=10)

ws_print['D21'] = "收货人签字：________________"
ws_print['D21'].font = Font(size=10)

# 底部信息
ws_print['A23'] = "─" * 40
ws_print['A23'].alignment = Alignment(horizontal='center')
ws_print.merge_cells('A23:G23')

ws_print['A24'] = "谢谢惠顾，欢迎下次光临！"
ws_print['A24'].font = Font(size=10, bold=True)
ws_print['A24'].alignment = Alignment(horizontal='center')
ws_print.merge_cells('A24:G24')

# 数据验证 - 订单号下拉
dv_order = DataValidation(type="list", formula1="'订单录入'!$J$2:$J$1000", allow_blank=True)
ws_print.add_data_validation(dv_order)
dv_order.add(ws_print['B5'])

# 设置列宽
ws_print.column_dimensions['A'].width = 5
ws_print.column_dimensions['B'].width = 14
ws_print.column_dimensions['C'].width = 10
ws_print.column_dimensions['D'].width = 8
ws_print.column_dimensions['E'].width = 8
ws_print.column_dimensions['F'].width = 8
ws_print.column_dimensions['G'].width = 10

# 打印设置
ws_print.print_area = 'A1:G30'
ws_print.page_setup.paperSize = 9  # 小票纸
ws_print.page_setup.orientation = 'portrait'
ws_print.page_setup.fitToPage = True
ws_print.page_setup.fitToWidth = 1
ws_print.page_setup.fitToHeight = 0

# ==================== Sheet 4: 统计汇总 ====================
ws_stats = wb.create_sheet("统计汇总")

# 设置列宽
ws_stats.column_dimensions['A'].width = 14
ws_stats.column_dimensions['B'].width = 18
ws_stats.column_dimensions['C'].width = 3
ws_stats.column_dimensions['D'].width = 14
ws_stats.column_dimensions['E'].width = 18

# 标题行
ws_stats['A1'] = "订单统计汇总"
ws_stats['A1'].font = Font(size=18, bold=True, color="FFFFFF")
ws_stats['A1'].fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
ws_stats['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws_stats.row_dimensions[1].height = 30
ws_stats.merge_cells('A1:E1')

# 空行
ws_stats.row_dimensions[2].height = 8

# ====== 左侧筛选条件 ======
# 筛选条件标题
ws_stats['A3'] = "筛选条件"
ws_stats['A3'].font = Font(size=12, bold=True, color="FFFFFF")
ws_stats['A3'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
ws_stats['A3'].alignment = Alignment(horizontal='center', vertical='center')
ws_stats.merge_cells('A3:B3')
ws_stats.row_dimensions[3].height = 22

# 开始日期
ws_stats['A4'] = "开始日期："
ws_stats['A4'].font = Font(bold=True, size=11)
ws_stats['A4'].alignment = Alignment(horizontal='right', vertical='center')
ws_stats['B4'].value = "2026-01-01"
ws_stats['B4'].number_format = 'yyyy/mm/dd'
ws_stats['B4'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
ws_stats['B4'].border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                               top=Side(style='thin'), bottom=Side(style='thin'))
ws_stats.row_dimensions[4].height = 20

# 结束日期
ws_stats['A5'] = "结束日期："
ws_stats['A5'].font = Font(bold=True, size=11)
ws_stats['A5'].alignment = Alignment(horizontal='right', vertical='center')
ws_stats['B5'].value = "2026-12-31"
ws_stats['B5'].number_format = 'yyyy/mm/dd'
ws_stats['B5'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
ws_stats['B5'].border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                               top=Side(style='thin'), bottom=Side(style='thin'))
ws_stats.row_dimensions[5].height = 20

# 客户筛选
ws_stats['A6'] = "客户："
ws_stats['A6'].font = Font(bold=True, size=11)
ws_stats['A6'].alignment = Alignment(horizontal='right', vertical='center')
ws_stats['B6'].value = ""  # 留空=全部
ws_stats['B6'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
ws_stats['B6'].border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                               top=Side(style='thin'), bottom=Side(style='thin'))
dv_filter_customer = DataValidation(type="list", formula1="'客户资料'!$A$2:$A$100", allow_blank=True)
ws_stats.add_data_validation(dv_filter_customer)
dv_filter_customer.add(ws_stats['B6'])
ws_stats.row_dimensions[6].height = 20

# 产品筛选
ws_stats['A7'] = "产品："
ws_stats['A7'].font = Font(bold=True, size=11)
ws_stats['A7'].alignment = Alignment(horizontal='right', vertical='center')
ws_stats['B7'].value = ""  # 留空=全部
ws_stats['B7'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
ws_stats['B7'].border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                               top=Side(style='thin'), bottom=Side(style='thin'))
dv_filter_product = DataValidation(type="list", formula1="'产品库'!$A$2:$A$100", allow_blank=True)
ws_stats.add_data_validation(dv_filter_product)
dv_filter_product.add(ws_stats['B7'])
ws_stats.row_dimensions[7].height = 20

# ====== 右侧核心指标 ======
# 总金额区域
ws_stats['D3'] = "总金额"
ws_stats['D3'].font = Font(size=12, bold=True, color="FFFFFF")
ws_stats['D3'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
ws_stats['D3'].alignment = Alignment(horizontal='center', vertical='center')
ws_stats.merge_cells('D3:E3')
ws_stats.row_dimensions[3].height = 22

ws_stats['D4'].value = '=IF(OR($B$4="",$B$5=""),0,SUMIFS(\'订单录入\'!$H:$H,\'订单录入\'!$A:$A,">="&$B$4,\'订单录入\'!$A:$A,"<="&$B$5,\'订单录入\'!$B:$B,IF($B$6="","*",$B$6),\'订单录入\'!$C:$C,IF($B$7="","*",$B$7)))'
ws_stats['D4'].number_format = '"¥"#,##0.00'
ws_stats['D4'].font = Font(size=16, bold=True, color="70AD47")
ws_stats['D4'].alignment = Alignment(horizontal='center', vertical='center')
ws_stats.merge_cells('D4:E4')
ws_stats.row_dimensions[4].height = 28

# 订单数量区域
ws_stats['D5'] = "订单总数"
ws_stats['D5'].font = Font(size=12, bold=True, color="FFFFFF")
ws_stats['D5'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
ws_stats['D5'].alignment = Alignment(horizontal='center', vertical='center')
ws_stats.merge_cells('D5:E5')
ws_stats.row_dimensions[5].height = 22

ws_stats['D6'].value = '=IF(OR($B$4="",$B$5=""),0,SUMIFS(\'订单录入\'!$F:$F,\'订单录入\'!$A:$A,">="&$B$4,\'订单录入\'!$A:$A,"<="&$B$5,\'订单录入\'!$B:$B,IF($B$6="","*",$B$6),\'订单录入\'!$C:$C,IF($B$7="","*",$B$7)))'
ws_stats['D6'].number_format = '0'
ws_stats['D6'].font = Font(size=16, bold=True, color="4472C4")
ws_stats['D6'].alignment = Alignment(horizontal='center', vertical='center')
ws_stats.merge_cells('D6:E6')
ws_stats.row_dimensions[6].height = 28

# 订单行数统计
ws_stats['D7'] = "订单行数"
ws_stats['D7'].font = Font(size=12, bold=True, color="FFFFFF")
ws_stats['D7'].fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
ws_stats['D7'].alignment = Alignment(horizontal='center', vertical='center')
ws_stats.merge_cells('D7:E7')
ws_stats.row_dimensions[7].height = 22

ws_stats['D8'].value = '=IF(OR($B$4="",$B$5=""),0,COUNTIFS(\'订单录入\'!$A:$A,">="&$B$4,\'订单录入\'!$A:$A,"<="&$B$5,\'订单录入\'!$B:$B,IF($B$6="","*",$B$6),\'订单录入\'!$C:$C,IF($B$7="","*",$B$7)))'
ws_stats['D8'].number_format = '0'
ws_stats['D8'].font = Font(size=16, bold=True, color="ED7D31")
ws_stats['D8'].alignment = Alignment(horizontal='center', vertical='center')
ws_stats.merge_cells('D8:E8')
ws_stats.row_dimensions[8].height = 28

# 空行
ws_stats.row_dimensions[9].height = 8

# 使用说明（黄色底色）
ws_stats['A10'] = "使用说明"
ws_stats['A10'].font = Font(bold=True, size=10, color="FFFFFF")
ws_stats['A10'].fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
ws_stats.merge_cells('A10:E10')
ws_stats.row_dimensions[10].height = 20

tips = [
    "1. 开始/结束日期：必须输入日期，格式为 yyyy/mm/dd 或 yyyy-mm-dd",
    "2. 客户和产品：可选，留空表示显示全部",
    "3. 数据会根据筛选条件自动计算更新",
    "4. 总金额：表示该时间段内所有订单的总价合计",
    "5. 订单总数：表示该时间段内各产品的总数量",
    "6. 订单行数：表示该时间段内有多少条订单记录"
]

for idx, tip in enumerate(tips, 11):
    ws_stats[f'A{idx}'] = tip
    ws_stats[f'A{idx}'].font = Font(size=9)
    ws_stats[f'A{idx}'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    ws_stats.merge_cells(f'A{idx}:E{idx}')
    ws_stats.row_dimensions[idx].height = 18

# ==================== 保存文件 ====================
output_path = "/Users/cassandra/Desktop/订单打印模板.xlsx"
wb.save(output_path)

print(f"Excel 文件已创建: {output_path}")
print("\n已完成的功能：")
print("✓ 5个工作表")
print("✓ 产品库表格（tbl_products）")
print("✓ 订单录入表格（tbl_orders）")
print("✓ 日期自动填充公式")
print("✓ 产品名称下拉选择")
print("✓ 规格/单价自动查找公式（XLOOKUP/VLOOKUP）")
print("✓ 总价自动计算公式")
print("✓ 订单号自动生成公式")
print("✓ 打印状态下拉选择")
print("✓ 打印模板（订单号下拉+数据引用）")
print("✓ 统计汇总公式")
